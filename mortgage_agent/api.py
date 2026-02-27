from __future__ import annotations

import os
from datetime import date
from io import BytesIO
from typing import Optional
import zipfile
from urllib.parse import quote
import calendar

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, HTTPException, Request, Depends
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel, Field, field_validator, model_validator
from slowapi import Limiter
from slowapi.errors import RateLimitExceeded
from slowapi.middleware import SlowAPIMiddleware
from slowapi.util import get_remote_address

from mortgage_agent.calculator import LoanParams, Prepayment, ScheduleRow, build_schedule, monthly_rate, normalize_method, simulate, simulate_recurring_extra, simulate_annual_recurring_extra
from mortgage_agent.report import generate_pdf


API_KEY = os.getenv("API_KEY")

DEFAULT_RATE_LIMIT = os.getenv("RATE_LIMIT_DEFAULT", "60/minute")
EXPORT_RATE_LIMIT = os.getenv("RATE_LIMIT_EXPORT", "15/minute")
MAX_TERM_MONTHS = int(os.getenv("MAX_TERM_MONTHS", "600"))
MAX_PRINCIPAL = float(os.getenv("MAX_PRINCIPAL", "30000000"))
MAX_ANNUAL_RATE = float(os.getenv("MAX_ANNUAL_RATE", "30"))
MAX_PREPAY_RATIO = float(os.getenv("MAX_PREPAY_RATIO", "1.0"))
MAX_SCHEDULE_ROWS = int(os.getenv("MAX_SCHEDULE_ROWS", "2000"))
MAX_EXPORT_BYTES = int(os.getenv("MAX_EXPORT_BYTES", str(6 * 1024 * 1024)))
ALLOWED_METHODS = {"equal_payment", "equal_principal"}


def _client_ip(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for")
    if forwarded:
        candidate = forwarded.split(",")[0].strip()
        if candidate:
            return candidate
    real_ip = request.headers.get("x-real-ip")
    if real_ip:
        return real_ip
    return get_remote_address(request)


limiter = Limiter(key_func=_client_ip, default_limits=[DEFAULT_RATE_LIMIT])

app = FastAPI(
    title="息策 Agent",
    description="不只是计算器，是帮你省下一辆车的房贷管家。",
    version="0.1.0",
)
app.state.limiter = limiter
app.add_middleware(SlowAPIMiddleware)


def require_api_key(request: Request):
    if not API_KEY:
        return
    provided = request.headers.get("x-api-key")
    if not provided or provided != API_KEY:
        raise HTTPException(status_code=401, detail="invalid or missing api key")


@app.exception_handler(RateLimitExceeded)
async def _rate_limit_handler(request: Request, exc: RateLimitExceeded):
    return JSONResponse(status_code=429, content={"detail": "Rate limit exceeded"})


class LoanRequest(BaseModel):
    # 基础贷款信息
    principal: float = Field(..., gt=0, le=MAX_PRINCIPAL, description="贷款总额/本金（元）")
    annual_rate: float = Field(..., ge=0, le=MAX_ANNUAL_RATE, description="年利率百分比，例如 3.5")
    term_months: int = Field(..., gt=0, le=MAX_TERM_MONTHS, description="贷款总期数（月），例如 360")
    method: str = Field(..., description="还款方式：equal_payment(等额本息) / equal_principal(等额本金)")

    # 已还信息（二选一：paid_periods 或 first_payment_date）
    paid_periods: Optional[int] = Field(0, ge=0, description="已还期数（月）。如传 None 则使用 first_payment_date 推算")
    first_payment_date: Optional[date] = Field(None, description="首次还款日期，用于推算已还期数")

    # 提前还款意向
    prepay_amount: float = Field(..., ge=0, description="本次计划提前还款金额（元）")

    # 可选参数
    invest_annual_rate: Optional[float] = Field(None, ge=0, le=MAX_ANNUAL_RATE, description="可选：理财年化收益率百分比")

    @field_validator("method")
    @classmethod
    def _validate_method(cls, value: str) -> str:
        if value not in ALLOWED_METHODS:
            raise ValueError(f"method must be one of {sorted(ALLOWED_METHODS)}")
        return value

    @model_validator(mode="after")
    def _validate_ranges(self) -> "LoanRequest":
        if self.paid_periods and self.paid_periods > self.term_months:
            raise ValueError("paid_periods cannot exceed term_months")
        if self.prepay_amount > self.principal * MAX_PREPAY_RATIO:
            raise ValueError("prepay_amount exceeds allowed ratio of principal")
        return self


class CalcResponse(BaseModel):
    # 仅返回：缩短年限方案 & 减少月供方案的节省利息
    savings_shorten_interest: float
    savings_reduce_payment_interest: float


class CombinedLoanRequest(BaseModel):
    fund_principal: float = Field(..., ge=0, le=MAX_PRINCIPAL, description="公积金贷款本金（元）；0 表示无公积金贷款")
    fund_annual_rate: float = Field(..., ge=0, le=MAX_ANNUAL_RATE, description="公积金贷款年利率（%）")
    commercial_principal: float = Field(..., ge=0, le=MAX_PRINCIPAL, description="商业贷款本金（元）；0 表示无商业贷款")
    commercial_annual_rate: float = Field(..., ge=0, le=MAX_ANNUAL_RATE, description="商业贷款年利率（%）")
    term_months: int = Field(..., gt=0, le=MAX_TERM_MONTHS, description="贷款总期数（月）")
    method: str = Field("equal_payment", description="还款方式：equal_payment(等额本息) / equal_principal(等额本金)")

    @field_validator("method")
    @classmethod
    def _validate_method(cls, value: str) -> str:
        if value not in ALLOWED_METHODS:
            raise ValueError(f"method must be one of {sorted(ALLOWED_METHODS)}")
        return value

    @model_validator(mode="after")
    def _validate_pairs(self) -> "CombinedLoanRequest":
        if self.fund_principal <= 0 and self.commercial_principal <= 0:
            raise ValueError("fund_principal 与 commercial_principal 不能同时为 0")
        if self.fund_principal > 0 and self.fund_annual_rate <= 0:
            raise ValueError("fund_annual_rate must be greater than 0 when fund_principal > 0")
        if self.commercial_principal > 0 and self.commercial_annual_rate <= 0:
            raise ValueError("commercial_annual_rate must be greater than 0 when commercial_principal > 0")
        return self


class RecurringInvestmentRequest(BaseModel):
    # 定投式还款：在原月供基础上追加固定金额
    principal: float = Field(..., gt=0, le=MAX_PRINCIPAL, description="贷款总额/本金（元）")
    annual_rate: float = Field(..., ge=0, le=MAX_ANNUAL_RATE, description="年利率百分比，例如 3.5")
    term_months: int = Field(..., gt=0, le=MAX_TERM_MONTHS, description="贷款总期数（月），例如 360")
    method: str = Field(..., description="还款方式：equal_payment(等额本息) / equal_principal(等额本金)")

    paid_periods: Optional[int] = Field(0, ge=0, description="已还期数（月）。如传 None 则使用 first_payment_date 推算")
    first_payment_date: Optional[date] = Field(None, description="首次还款日期，用于推算已还期数")

    recurring_extra_amount: float = Field(..., gt=0, description="每月固定追加的定投还款额（元）")
    recurring_day: Optional[int] = Field(None, ge=1, le=28, description="每月定投日(1-28)，仅用于提示")

    @field_validator("method")
    @classmethod
    def _validate_method(cls, value: str) -> str:
        if value not in ALLOWED_METHODS:
            raise ValueError(f"method must be one of {sorted(ALLOWED_METHODS)}")
        return value

    @model_validator(mode="after")
    def _validate_ranges(self) -> "RecurringInvestmentRequest":
        if self.paid_periods and self.paid_periods > self.term_months:
            raise ValueError("paid_periods cannot exceed term_months")
        if self.recurring_extra_amount > self.principal * MAX_PREPAY_RATIO:
            raise ValueError("recurring_extra_amount exceeds allowed ratio of principal")
        return self


class RecurringCalcResponse(BaseModel):
    months_to_payoff: int
    total_interest_with_recurring: float
    base_total_interest: float
    interest_savings_vs_base: float
    base_remaining_interest: float
    total_payment_with_recurring: float
    base_monthly_payment: float
    recurring_extra_amount: float


class ScheduledRecurringRequest(BaseModel):
    principal: float = Field(..., gt=0, le=MAX_PRINCIPAL, description="贷款总额/本金（元）")
    annual_rate: float = Field(..., ge=0, le=MAX_ANNUAL_RATE, description="年利率百分比，例如 3.5")
    term_months: int = Field(..., gt=0, le=MAX_TERM_MONTHS, description="贷款总期数（月），例如 360")
    method: str = Field(..., description="还款方式：equal_payment(等额本息) / equal_principal(等额本金)")
    paid_periods: Optional[int] = Field(None, ge=0, description="可选：已还期数（月）；如不填结合 first_payment_date + as_of_date 推算")
    first_payment_date: Optional[date] = Field(None, description="可选：首次还款日期，用于推算已还期数")
    as_of_date: Optional[date] = Field(None, description="测算日期/本次还款日，用于计算已还期数")
    recurring_extra_amount: float = Field(..., gt=0, description="每月固定追加的定投还款额（元）")
    recurring_day: Optional[int] = Field(None, ge=1, le=28, description="每月定投日(1-28)，仅用于提示")
    recurring_start_date: Optional[date] = Field(None, description="定投开始日期，未填则默认从 as_of_date 开始生效")
    frequency: str = Field("monthly", description="定投频率：monthly(默认)/annual")
    annual_month: Optional[int] = Field(None, ge=1, le=12, description="frequency=annual 时：每年追加还款的月份(1-12)")
    annual_day: Optional[int] = Field(None, ge=1, le=31, description="frequency=annual 时：每年追加还款的日期(1-31)")

    @field_validator("method")
    @classmethod
    def _validate_method(cls, value: str) -> str:
        if value not in ALLOWED_METHODS:
            raise ValueError(f"method must be one of {sorted(ALLOWED_METHODS)}")
        return value

    @field_validator("frequency")
    @classmethod
    def _validate_frequency(cls, value: str) -> str:
        normalized = (value or "").lower()
        if normalized not in {"monthly", "annual"}:
            raise ValueError("frequency must be monthly or annual")
        return normalized

    @model_validator(mode="after")
    def _validate_ranges(self) -> "ScheduledRecurringRequest":
        if self.paid_periods and self.paid_periods > self.term_months:
            raise ValueError("paid_periods cannot exceed term_months")
        if self.recurring_extra_amount > self.principal * MAX_PREPAY_RATIO:
            raise ValueError("recurring_extra_amount exceeds allowed ratio of principal")
        if self.first_payment_date and self.as_of_date and self.as_of_date < self.first_payment_date:
            raise ValueError("as_of_date cannot be earlier than first_payment_date")
        if self.frequency == "annual":
            if not self.recurring_start_date and (not self.annual_month or not self.annual_day):
                raise ValueError("annual_month and annual_day are required when frequency=annual without recurring_start_date")
        return self


class ScheduledRecurringResponse(BaseModel):
    months_to_payoff: int
    payoff_date: Optional[date]
    total_interest_with_recurring: float
    base_total_interest: float
    interest_savings_vs_base: float
    base_remaining_interest: float
    total_payment_with_recurring: float
    base_monthly_payment: float
    recurring_extra_amount: float
    start_offset_months: int
    as_of_date: Optional[date]
    recurring_start_date: Optional[date]
    first_annual_extra_date: Optional[date]


class LumpSumRecurringRequest(BaseModel):
    principal: float = Field(..., gt=0, le=MAX_PRINCIPAL, description="贷款总额/本金（元）")
    annual_rate: float = Field(..., ge=0, le=MAX_ANNUAL_RATE, description="年利率百分比，例如 3.5")
    term_months: int = Field(..., gt=0, le=MAX_TERM_MONTHS, description="贷款总期数（月），例如 360")
    method: str = Field(..., description="还款方式：equal_payment(等额本息) / equal_principal(等额本金)")
    paid_periods: Optional[int] = Field(None, ge=0, description="已还期数（月）；如不填可结合 first_payment_date 推算")
    first_payment_date: Optional[date] = Field(None, description="首次还款日期，用于推算已还期数")
    repayment_date: Optional[date] = Field(None, description="本次测算/还款日，用于确定已还期数及预计结清日")
    recurring_extra_amount: float = Field(..., gt=0, description="每期固定追加的定投还款额（元）")
    recurring_day: Optional[int] = Field(None, ge=1, le=28, description="每月定投日(1-28)，用于提示")
    recurring_start_date: Optional[date] = Field(None, description="定投开始日期；为空则默认从 repayment_date 开始")

    @field_validator("method")
    @classmethod
    def _validate_method(cls, value: str) -> str:
        if value not in ALLOWED_METHODS:
            raise ValueError(f"method must be one of {sorted(ALLOWED_METHODS)}")
        return value

    @model_validator(mode="after")
    def _validate_ranges(self) -> "LumpSumRecurringRequest":
        if self.paid_periods and self.paid_periods > self.term_months:
            raise ValueError("paid_periods cannot exceed term_months")
        if self.recurring_extra_amount > self.principal * MAX_PREPAY_RATIO:
            raise ValueError("recurring_extra_amount exceeds allowed ratio of principal")
        if self.first_payment_date and self.repayment_date and self.repayment_date < self.first_payment_date:
            raise ValueError("repayment_date cannot be earlier than first_payment_date")
        return self


class LumpSumRecurringResponse(BaseModel):
    months_to_payoff: int
    payoff_date: Optional[date]
    total_interest_with_recurring: float
    base_total_interest: float
    interest_savings_vs_base: float
    base_remaining_interest: float
    total_payment_with_recurring: float
    base_monthly_payment: float
    recurring_extra_amount: float
    start_offset_months: int
    repayment_date: Optional[date]
    recurring_start_date: Optional[date]


class AnnualRecurringRequest(BaseModel):
    principal: float = Field(..., gt=0, le=MAX_PRINCIPAL, description="贷款总额/本金（元）")
    annual_rate: float = Field(..., ge=0, le=MAX_ANNUAL_RATE, description="年利率百分比，例如 3.5")
    term_months: int = Field(..., gt=0, le=MAX_TERM_MONTHS, description="贷款总期数（月），例如 360")
    method: str = Field(..., description="还款方式：equal_payment(等额本息) / equal_principal(等额本金)")
    paid_periods: Optional[int] = Field(None, ge=0, description="已还期数（月）；如不填结合 first_payment_date + as_of_date 推算")
    first_payment_date: Optional[date] = Field(None, description="首次还款日期，用于推算已还期数")
    as_of_date: Optional[date] = Field(None, description="测算日期/本次还款日，用于计算已还期数")
    annual_extra_amount: float = Field(..., gt=0, description="每年固定追加的定投还款额（元）")
    recurring_start_date: date = Field(..., description="首次年度定投日期（例如 2022-12-29）；之后每年同日定投")

    @field_validator("method")
    @classmethod
    def _validate_method(cls, value: str) -> str:
        if value not in ALLOWED_METHODS:
            raise ValueError(f"method must be one of {sorted(ALLOWED_METHODS)}")
        return value

    @model_validator(mode="after")
    def _validate_ranges(self) -> "AnnualRecurringRequest":
        if self.paid_periods and self.paid_periods > self.term_months:
            raise ValueError("paid_periods cannot exceed term_months")
        if self.annual_extra_amount > self.principal * MAX_PREPAY_RATIO:
            raise ValueError("annual_extra_amount exceeds allowed ratio of principal")
        if self.first_payment_date and self.as_of_date and self.as_of_date < self.first_payment_date:
            raise ValueError("as_of_date cannot be earlier than first_payment_date")
        return self


class AnnualRecurringResponse(BaseModel):
    months_to_payoff: int
    payoff_date: Optional[date]
    total_interest_with_recurring: float
    base_total_interest: float
    interest_savings_vs_base: float
    base_remaining_interest: float
    total_payment_with_recurring: float
    base_monthly_payment: float
    annual_extra_amount: float
    start_offset_months: int
    as_of_date: Optional[date]
    first_annual_extra_date: Optional[date]


@app.get("/health", tags=["health"])
@limiter.exempt
def health() -> dict:
    return {"status": "ok"}


@app.post(
    "/v1/mortgages/prepayment:calc",
    tags=["mortgage"],
    responses={400: {"description": "Invalid loan or prepayment parameters"}},
)
@limiter.limit(DEFAULT_RATE_LIMIT)
def calc_prepayment(request: Request, body: LoanRequest, _=Depends(require_api_key)) -> CalcResponse:
    try:
        params = LoanParams(
            principal=body.principal,
            annual_rate=body.annual_rate,
            term_months=body.term_months,
            method=body.method,
            paid_periods=body.paid_periods,  # 允许 None
            first_payment_date=body.first_payment_date,
        )
        prepay = Prepayment(amount=body.prepay_amount, invest_annual_rate=body.invest_annual_rate)
        result = simulate(params, prepay)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return CalcResponse(
        savings_shorten_interest=float(result.savings_shorten),
        savings_reduce_payment_interest=float(result.savings_reduce),
    )

@app.post(
    "/v1/mortgages/prepayment:export-zip",
    tags=["mortgage"],
    responses={400: {"description": "Invalid loan or prepayment parameters"}},
)
@limiter.limit(EXPORT_RATE_LIMIT)
def export_zip(request: Request, body: LoanRequest, _=Depends(require_api_key)):
    """导出还款明细 ZIP（原方案/减少月供/缩短年限，各一份 Excel）。"""
    try:
        params = LoanParams(
            principal=body.principal,
            annual_rate=body.annual_rate,
            term_months=body.term_months,
            method=body.method,
            paid_periods=body.paid_periods,
            first_payment_date=body.first_payment_date,
        )
        prepay = Prepayment(amount=body.prepay_amount, invest_annual_rate=body.invest_annual_rate)
        result = simulate(params, prepay)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    _ensure_row_limit(len(result.base_schedule), "base_schedule")
    _ensure_row_limit(len(result.reduced_schedule), "reduced_schedule")
    _ensure_row_limit(len(result.shorten_schedule), "shorten_schedule")

    pdf_bytes = generate_pdf(
        result=result,
        prepayment=prepay,
        original_principal=body.principal,
        original_annual_rate=body.annual_rate,
        original_term_months=body.term_months,
        original_method=body.method,
    )

    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("原方案月供明细.xlsx", _schedule_to_xlsx(result.base_schedule))
        zf.writestr("提前还款-减少月供-月供明细.xlsx", _schedule_to_xlsx(result.reduced_schedule))
        zf.writestr("提前还款-缩短年限-月供明细.xlsx", _schedule_to_xlsx(result.shorten_schedule))
        zf.writestr("提前还款-分析报告.pdf", pdf_bytes)
    zip_buf.seek(0)

    zip_bytes = zip_buf.getvalue()
    _ensure_export_size(len(zip_bytes))

    return StreamingResponse(
        BytesIO(zip_bytes),
        media_type="application/zip",
        headers={
            "Content-Disposition": "attachment; filename=提前还款分析报告.zip",
            "X-Savings-Reduce": f"{float(result.savings_reduce):.2f}",
            "X-Savings-Shorten": f"{float(result.savings_shorten):.2f}",
        },
    )

@app.post(
    "/v1/mortgages/combined:export-xlsx",
    tags=["mortgage"],
    responses={400: {"description": "Invalid loan parameters"}},
)
@limiter.limit(EXPORT_RATE_LIMIT)
def export_combined_schedule(request: Request, body: CombinedLoanRequest, _=Depends(require_api_key)):
    """组合贷（公积金 + 商贷）还款计划导出 Excel，响应头返回总利息。"""
    include_fund = body.fund_principal > 0
    include_commercial = body.commercial_principal > 0

    try:
        method = normalize_method(body.method)
        fund_schedule = build_schedule(
            body.fund_principal,
            monthly_rate(body.fund_annual_rate),
            body.term_months,
            method,
        ) if include_fund else []
        commercial_schedule = build_schedule(
            body.commercial_principal,
            monthly_rate(body.commercial_annual_rate),
            body.term_months,
            method,
        ) if include_commercial else []
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    _ensure_row_limit(max(len(fund_schedule), len(commercial_schedule)), "combined_schedule")

    max_len = max(len(fund_schedule), len(commercial_schedule))
    combined: list[ScheduleRow] = []
    total_interest = 0.0

    for idx in range(max_len):
        fund = fund_schedule[idx] if idx < len(fund_schedule) else ScheduleRow(idx + 1, 0, 0, 0, 0)
        commercial = commercial_schedule[idx] if idx < len(commercial_schedule) else ScheduleRow(idx + 1, 0, 0, 0, 0)
        payment = fund.payment + commercial.payment
        principal = fund.principal + commercial.principal
        interest = fund.interest + commercial.interest
        balance = fund.balance + commercial.balance
        combined.append(ScheduleRow(idx + 1, payment, principal, interest, balance))
        total_interest += interest

    zip_bytes = _combined_schedule_to_xlsx(
        combined,
        commercial_schedule,
        fund_schedule,
        include_commercial=include_commercial,
        include_fund=include_fund,
    )

    _ensure_export_size(len(zip_bytes))

    return StreamingResponse(
        BytesIO(zip_bytes),
        media_type="application/zip",
        headers={
            "Content-Disposition": "attachment; filename=loan_schedules.zip; "
            f"filename*=UTF-8''{quote('房贷月供明细.zip')}",
            "X-Total-Interest": f"{float(total_interest):.2f}",
        },
    )

@app.post(
    "/v1/mortgages/recurring:calc",
    tags=["mortgage"],
    responses={400: {"description": "Invalid recurring repayment parameters"}},
)
@limiter.limit(DEFAULT_RATE_LIMIT)
def calc_recurring_extra(request: Request, body: RecurringInvestmentRequest, _=Depends(require_api_key)) -> RecurringCalcResponse:
    try:
        params = LoanParams(
            principal=body.principal,
            annual_rate=body.annual_rate,
            term_months=body.term_months,
            method=body.method,
            paid_periods=body.paid_periods,
            first_payment_date=body.first_payment_date,
        )
        result = simulate_recurring_extra(params, recurring_extra=body.recurring_extra_amount)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return RecurringCalcResponse(
        months_to_payoff=result.months_with_recurring,
        total_interest_with_recurring=float(result.total_interest_with_recurring),
        base_total_interest=float(result.base_total_interest),
        interest_savings_vs_base=float(result.interest_savings),
        base_remaining_interest=float(result.base_remaining_interest),
        total_payment_with_recurring=float(result.total_payment_with_recurring),
        base_monthly_payment=float(result.base_monthly_payment),
        recurring_extra_amount=float(result.recurring_extra_payment),
    )


@app.post(
    "/v1/mortgages/recurring:calc-scheduled",
    tags=["mortgage"],
    responses={400: {"description": "Invalid scheduled recurring repayment parameters"}},
)
@limiter.limit(DEFAULT_RATE_LIMIT)
def calc_scheduled_recurring(request: Request, body: ScheduledRecurringRequest, _=Depends(require_api_key)) -> ScheduledRecurringResponse:
    try:
        params = LoanParams(
            principal=body.principal,
            annual_rate=body.annual_rate,
            term_months=body.term_months,
            method=body.method,
            paid_periods=body.paid_periods,
            first_payment_date=body.first_payment_date,
        )
        as_of_date = body.as_of_date or body.first_payment_date or date.today()
        start_offset = 0
        first_extra_date: Optional[date] = None
        if body.frequency == "annual":
            first_extra_date = _resolve_first_annual_date(
                as_of_date,
                recurring_start_date=body.recurring_start_date,
                annual_month=body.annual_month,
                annual_day=body.annual_day,
            )
            start_offset = _months_between(as_of_date, first_extra_date)
            result = simulate_annual_recurring_extra(
                params,
                annual_extra=body.recurring_extra_amount,
                as_of_date=as_of_date,
                months_until_first_extra=start_offset,
            )
        else:
            if body.recurring_start_date:
                start_offset = _months_between(as_of_date, body.recurring_start_date)
            result = simulate_recurring_extra(
                params,
                recurring_extra=body.recurring_extra_amount,
                as_of_date=as_of_date,
                start_offset_months=start_offset,
            )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    payoff_date = _add_months(as_of_date, result.months_with_recurring) if as_of_date else None

    return ScheduledRecurringResponse(
        months_to_payoff=result.months_with_recurring,
        payoff_date=payoff_date,
        total_interest_with_recurring=float(result.total_interest_with_recurring),
        base_total_interest=float(result.base_total_interest),
        interest_savings_vs_base=float(result.interest_savings),
        base_remaining_interest=float(result.base_remaining_interest),
        total_payment_with_recurring=float(result.total_payment_with_recurring),
        base_monthly_payment=float(result.base_monthly_payment),
        recurring_extra_amount=float(result.recurring_extra_payment),
        start_offset_months=start_offset,
        as_of_date=as_of_date,
        recurring_start_date=body.recurring_start_date,
        first_annual_extra_date=first_extra_date,
    )


@app.post(
    "/v1/mortgages/recurring:lump-sum",
    tags=["mortgage"],
    responses={400: {"description": "Invalid lump-sum recurring repayment parameters"}},
)
@limiter.limit(DEFAULT_RATE_LIMIT)
def calc_lump_sum_recurring(request: Request, body: LumpSumRecurringRequest, _=Depends(require_api_key)) -> LumpSumRecurringResponse:
    try:
        params = LoanParams(
            principal=body.principal,
            annual_rate=body.annual_rate,
            term_months=body.term_months,
            method=body.method,
            paid_periods=body.paid_periods,
            first_payment_date=body.first_payment_date,
        )
        repayment_date = body.repayment_date or body.first_payment_date or date.today()
        start_offset = 0
        if body.recurring_start_date:
            start_offset = _months_between(repayment_date, body.recurring_start_date)
        result = simulate_recurring_extra(
            params,
            recurring_extra=body.recurring_extra_amount,
            as_of_date=repayment_date,
            start_offset_months=start_offset,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    payoff_date = _add_months(repayment_date, result.months_with_recurring) if repayment_date else None

    return LumpSumRecurringResponse(
        months_to_payoff=result.months_with_recurring,
        payoff_date=payoff_date,
        total_interest_with_recurring=float(result.total_interest_with_recurring),
        base_total_interest=float(result.base_total_interest),
        interest_savings_vs_base=float(result.interest_savings),
        base_remaining_interest=float(result.base_remaining_interest),
        total_payment_with_recurring=float(result.total_payment_with_recurring),
        base_monthly_payment=float(result.base_monthly_payment),
        recurring_extra_amount=float(result.recurring_extra_payment),
        start_offset_months=start_offset,
        repayment_date=repayment_date,
        recurring_start_date=body.recurring_start_date,
    )


@app.post(
    "/v1/mortgages/recurring:annual",
    tags=["mortgage"],
    responses={400: {"description": "Invalid annual recurring repayment parameters"}},
)
@limiter.limit(DEFAULT_RATE_LIMIT)
def calc_annual_recurring(request: Request, body: AnnualRecurringRequest, _=Depends(require_api_key)) -> AnnualRecurringResponse:
    try:
        params = LoanParams(
            principal=body.principal,
            annual_rate=body.annual_rate,
            term_months=body.term_months,
            method=body.method,
            paid_periods=body.paid_periods,
            first_payment_date=body.first_payment_date,
        )
        as_of_date = body.as_of_date or body.first_payment_date or date.today()
        first_extra_date = _resolve_first_annual_date(
            as_of_date,
            recurring_start_date=body.recurring_start_date,
            annual_month=None,
            annual_day=None,
        )
        start_offset = _months_between(as_of_date, first_extra_date)
        result = simulate_annual_recurring_extra(
            params,
            annual_extra=body.annual_extra_amount,
            as_of_date=as_of_date,
            months_until_first_extra=start_offset,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    payoff_date = _add_months(as_of_date, result.months_with_recurring) if as_of_date else None

    return AnnualRecurringResponse(
        months_to_payoff=result.months_with_recurring,
        payoff_date=payoff_date,
        total_interest_with_recurring=float(result.total_interest_with_recurring),
        base_total_interest=float(result.base_total_interest),
        interest_savings_vs_base=float(result.interest_savings),
        base_remaining_interest=float(result.base_remaining_interest),
        total_payment_with_recurring=float(result.total_payment_with_recurring),
        base_monthly_payment=float(result.base_monthly_payment),
        annual_extra_amount=float(result.recurring_extra_payment),
        start_offset_months=start_offset,
        as_of_date=as_of_date,
        first_annual_extra_date=first_extra_date,
    )


def _months_between(start: date, end: date) -> int:
    if end <= start:
        return 0
    months = (end.year - start.year) * 12 + (end.month - start.month)
    if end.day < start.day:
        months -= 1
    return max(months, 0)


def _next_annual_date(as_of: date, month: int, day: int) -> date:
    # 返回 as_of 之后（含当月）的下一个“month/day”日期。
    target_year = as_of.year
    day = min(day, calendar.monthrange(target_year, month)[1])
    candidate = date(target_year, month, day)
    if candidate < as_of:
        target_year += 1
        day = min(day, calendar.monthrange(target_year, month)[1])
        candidate = date(target_year, month, day)
    return candidate


def _add_months(src: date, months: int) -> date:
    month = src.month - 1 + months
    year = src.year + month // 12
    month = month % 12 + 1
    day = min(src.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def _schedule_to_xlsx(schedule) -> bytes:
    """将还款计划列表导出为 Excel（xlsx），返回二进制。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    headers = ["期数", "月供", "本金", "利息", "余额", "利息占比"]
    ws.append(headers)

    header_font = Font(bold=True, name="Arial", size=11, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    header_fill = PatternFill("solid", fgColor="0F172A")
    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    border = Border(bottom=Side(style="thin", color="E2E8F0"))
    align_right = Alignment(horizontal="right")
    align_center = Alignment(horizontal="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    for idx, row in enumerate(schedule, start=2):
        ratio = (row.interest / row.payment) if row.payment else 0.0
        ws.append([
            row.month_index,
            round(row.payment, 2),
            round(row.principal, 2),
            round(row.interest, 2),
            round(row.balance, 2),
            f"{ratio*100:.2f}%",
        ])
        for col_idx in range(1, 7):
            cell = ws.cell(row=idx, column=col_idx)
            cell.font = body_font
            cell.alignment = align_right if col_idx > 1 else align_center
            if idx % 2 == 0:
                cell.fill = alt_fill
            cell.border = border
        # 利息占比<50% 标红
        ratio_cell = ws.cell(row=idx, column=6)
        if ratio < 0.5:
            ratio_cell.font = Font(name="Arial", size=10, color="EF4444")

    # 列宽
    widths = [8, 14, 14, 14, 16, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _combined_schedule_to_xlsx(
    combined: list[ScheduleRow],
    commercial_schedule: list[ScheduleRow],
    fund_schedule: list[ScheduleRow],
    include_commercial: bool,
    include_fund: bool,
) -> bytes:
    """组合贷专用：动态输出商贷/公积金列，含各自利息占比与总利息占比。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined"

    headers = ["期数", "月供总额"]
    if include_commercial:
        headers += [
            "商贷月供总额",
            "商贷本金",
            "商贷利息",
            "商贷余额",
            "商贷利息占比",
        ]
    if include_fund:
        headers += [
            "公积金月供总额",
            "公积金本金",
            "公积金利息",
            "公积金余额",
            "公积金利息占比",
        ]
    headers.append("利息总占比")
    ws.append(headers)

    header_font = Font(bold=True, name="Arial", size=11, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    header_fill_base = PatternFill("solid", fgColor="0F172A")
    header_fill_commercial = PatternFill("solid", fgColor="1D4ED8")
    header_fill_fund = PatternFill("solid", fgColor="047857")
    body_fill_commercial = PatternFill("solid", fgColor="EFF6FF")
    body_fill_fund = PatternFill("solid", fgColor="ECFDF3")
    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    border = Border(bottom=Side(style="thin", color="E2E8F0"))
    align_right = Alignment(horizontal="right")
    align_center = Alignment(horizontal="center")

    commercial_cols = [idx for idx, h in enumerate(headers, start=1) if h.startswith("商贷")]
    fund_cols = [idx for idx, h in enumerate(headers, start=1) if h.startswith("公积金")]

    for idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        if idx in commercial_cols:
            cell.fill = header_fill_commercial
        elif idx in fund_cols:
            cell.fill = header_fill_fund
        else:
            cell.fill = header_fill_base
        cell.alignment = align_center

    max_len = len(combined)
    for idx in range(max_len):
        combined_row = combined[idx]
        row_values = [
            combined_row.month_index,
            round(combined_row.payment, 2),
        ]

        if include_commercial:
            c = commercial_schedule[idx] if idx < len(commercial_schedule) else ScheduleRow(idx + 1, 0, 0, 0, 0)
            c_ratio = (c.interest / c.payment * 100) if c.payment else 0.0
            row_values += [
                round(c.payment, 2),
                round(c.principal, 2),
                round(c.interest, 2),
                round(c.balance, 2),
                f"{c_ratio:.2f}%",
            ]

        if include_fund:
            f = fund_schedule[idx] if idx < len(fund_schedule) else ScheduleRow(idx + 1, 0, 0, 0, 0)
            f_ratio = (f.interest / f.payment * 100) if f.payment else 0.0
            row_values += [
                round(f.payment, 2),
                round(f.principal, 2),
                round(f.interest, 2),
                round(f.balance, 2),
                f"{f_ratio:.2f}%",
            ]

        total_ratio = (combined_row.interest / combined_row.payment * 100) if combined_row.payment else 0.0
        row_values.append(f"{total_ratio:.2f}%")

        ws.append(row_values)

        for col_idx in range(1, len(row_values) + 1):
            cell = ws.cell(row=idx + 2, column=col_idx)
            cell.font = body_font
            cell.alignment = align_right if col_idx > 1 else align_center
            if col_idx in commercial_cols:
                cell.fill = body_fill_commercial
            elif col_idx in fund_cols:
                cell.fill = body_fill_fund
            elif (idx + 2) % 2 == 0:
                cell.fill = alt_fill
            cell.border = border

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        # 将 xlsx 打包为 zip，便于前端统一处理
        zf.writestr("房贷月供明细.xlsx", buf.getvalue())
    zip_buf.seek(0)
    return zip_buf.getvalue()


def _ensure_row_limit(rows: int, label: str) -> None:
    if rows > MAX_SCHEDULE_ROWS:
        raise HTTPException(status_code=413, detail=f"{label} too large, exceeds {MAX_SCHEDULE_ROWS} rows limit")


def _ensure_export_size(size_bytes: int) -> None:
    if size_bytes > MAX_EXPORT_BYTES:
        raise HTTPException(status_code=413, detail="export file too large")


def _resolve_first_annual_date(
    as_of: date,
    *,
    recurring_start_date: Optional[date],
    annual_month: Optional[int],
    annual_day: Optional[int],
) -> date:
    if recurring_start_date:
        if recurring_start_date >= as_of:
            return recurring_start_date
        month = recurring_start_date.month
        day = recurring_start_date.day
    else:
        if not annual_month or not annual_day:
            raise ValueError("annual_month and annual_day are required when recurring_start_date is not provided")
        month = annual_month
        day = annual_day
    return _next_annual_date(as_of, month, day)

